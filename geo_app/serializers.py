import logging
from rest_framework import serializers
from openpyxl import load_workbook
from .models import Data


logger = logging.getLogger(__name__)


class DataSerializer(serializers.ModelSerializer):
    class Meta:
        model = Data
        fields = '__all__'


class FileUploadSerializer(serializers.Serializer):
    file = serializers.FileField()

    def validate_file(self, value):
        if not value.name.endswith('.xlsx'):
            raise serializers.ValidationError("Файл должен быть в формате .xlsx")
        return value

    def process_file(self, file):
        wb = load_workbook(file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                ne = row[0]
                address = row[1]
                
                coordinates = row[2].split(", ")
                latitude = float(coordinates[0])
                longitude = float(coordinates[1])

                technologies = row[3].replace(" ", "").split(",")
                gsm = "gsm" in technologies
                umts = "umts" in technologies
                lte = "lte" in technologies

                status = int(row[4])

                Data.objects.create(
                    ne=ne,
                    address=address,
                    latitude=latitude,
                    longitude=longitude,
                    gsm=gsm,
                    umts=umts,
                    lte=lte,
                    status=status,
                )
            except Exception as e:
                # Логируем ошибку, но продолжаем обработку других строк
                logger.error(f"Ошибка при обработке строки: {row}, ошибка: {e}")
                continue
